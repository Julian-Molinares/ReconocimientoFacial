"""
MÓDULO 04 — Registro de Asistencia en Excel
Columnas: Nombre | Similitud (%) | Hora | Estado

- El archivo Excel se crea nuevo cada vez que se inicia main.py
- Un alumno se marca Presente solo si fue reconocido N veces seguidas
- Controla falsos positivos con contador de confirmaciones
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os


CARPETA_ASISTENCIAS = os.path.join(os.path.dirname(__file__), "asistencias")
COLUMNAS            = ["Nombre", "Similitud (%)", "Hora", "Estado"]

# Número de veces consecutivas que debe reconocerse al alumno
# antes de registrar su asistencia
CONFIRMACIONES_REQUERIDAS = 5


def obtener_ruta_excel() -> str:
    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return os.path.join(CARPETA_ASISTENCIAS, f"asistencia_{fecha}.xlsx")


def crear_excel_sesion() -> str:
    """
    Crea un archivo Excel NUEVO al iniciar cada sesión.
    El nombre incluye fecha y hora para no sobreescribir sesiones anteriores.
    Retorna la ruta del archivo creado.
    """
    os.makedirs(CARPETA_ASISTENCIAS, exist_ok=True)
    ruta = obtener_ruta_excel()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    fill = PatternFill("solid", fgColor="2E4057")
    font = Font(color="FFFFFF", bold=True)

    for col, nombre_col in enumerate(COLUMNAS, start=1):
        c           = ws.cell(row=1, column=col, value=nombre_col)
        c.fill      = fill
        c.font      = font
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15

    wb.save(ruta)
    print(f"[registro_excel] Nueva sesión: {ruta}")
    return ruta


def ya_presente(ruta: str, nombre: str) -> bool:
    """Verifica si el alumno ya fue marcado como Presente en esta sesión."""
    if not os.path.exists(ruta):
        return False
    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] == nombre and fila[3] == "Presente":
            return True
    return False


def obtener_fila_alumno(ws, nombre: str):
    """Retorna el número de fila del alumno o None si no existe."""
    for fila in ws.iter_rows(min_row=2):
        if fila[0].value == nombre:
            return fila[0].row
    return None


def distancia_a_porcentaje(distancia: float) -> float:
    """
    Convierte distancia euclidiana a porcentaje de similitud.
    Fórmula: similitud (%) = (1 - distancia / 2) * 100
    """
    porcentaje = (1.0 - distancia / 2.0) * 100.0
    return round(max(0.0, min(100.0, porcentaje)), 2)


def registrar_asistencia(ruta: str, nombre: str, distancia: float,
                         estado: str = "Presente") -> bool:
    """
    Registra o actualiza la asistencia en el Excel de la sesión actual.

    Args:
        ruta:      ruta del Excel de la sesión actual (de crear_excel_sesion)
        nombre:    nombre del alumno
        distancia: distancia euclidiana obtenida
        estado:    "Presente" o "No Presente"

    Returns:
        True si se registró | False si ya estaba presente
    """
    if ya_presente(ruta, nombre) and estado == "Presente":
        print(f"[registro_excel] {nombre} ya está Presente — ignorando")
        return False

    porcentaje = distancia_a_porcentaje(distancia)
    hora       = datetime.now().strftime("%H:%M:%S")

    wb       = openpyxl.load_workbook(ruta)
    ws       = wb.active
    num_fila = obtener_fila_alumno(ws, nombre)

    if num_fila and estado == "Presente":
        # Actualizar fila existente No Presente → Presente
        ws.cell(row=num_fila, column=2).value = f"{porcentaje}%"
        ws.cell(row=num_fila, column=3).value = hora
        ws.cell(row=num_fila, column=4).value = "Presente"
        color = "C8F7C5"
        print(f"[registro_excel] ↑ {nombre} actualizado a Presente | {porcentaje}% | {hora}")
    else:
        ws.append([nombre, f"{porcentaje}%" if estado == "Presente" else "—", hora, estado])
        num_fila = ws.max_row
        color    = "C8F7C5" if estado == "Presente" else "F7C5C5"
        print(f"[registro_excel] ✅ {nombre} | {porcentaje}% | {hora} | {estado}")

    fill = PatternFill("solid", fgColor=color)
    for col in range(1, 5):
        ws.cell(row=num_fila, column=col).fill      = fill
        ws.cell(row=num_fila, column=col).alignment = Alignment(horizontal="center")

    wb.save(ruta)
    return True


def marcar_ausentes(ruta: str, bd: dict):
    """
    Agrega 'No Presente' a los alumnos que no fueron detectados en la sesión.
    """
    wb          = openpyxl.load_workbook(ruta)
    ws          = wb.active
    ya_en_excel = set()
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0]:
            ya_en_excel.add(fila[0])

    ausentes = 0
    for nombre in bd.keys():
        if nombre not in ya_en_excel:
            registrar_asistencia(ruta, nombre, distancia=2.0, estado="No Presente")
            ausentes += 1

    print(f"[registro_excel] Ausentes marcados: {ausentes}")


def obtener_presentes(ruta: str) -> list:
    """Retorna la lista de alumnos con estado Presente en la sesión actual."""
    if not os.path.exists(ruta):
        return []
    wb        = openpyxl.load_workbook(ruta)
    ws        = wb.active
    presentes = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] and fila[3] == "Presente":
            presentes.append(fila[0])
    return presentes


# ──────────────────────────────────────────────────────────────
# CONTADOR DE CONFIRMACIONES
# Evita falsos positivos: solo registra después de N detecciones
# ──────────────────────────────────────────────────────────────
class ContadorConfirmaciones:
    """
    Lleva la cuenta de cuántas veces consecutivas fue reconocida
    cada persona. Solo registra la asistencia al llegar a N confirmaciones.

    Ejemplo con N=5:
      - Frame 1: Imanol detectado → contador Imanol = 1
      - Frame 2: Imanol detectado → contador Imanol = 2
      - Frame 3: Desconocido      → contador Imanol = 0 (reset)
      - Frame 4: Imanol detectado → contador Imanol = 1
      ...
      - Frame N: Imanol detectado → contador = 5 → REGISTRAR ✅
    """
    def __init__(self, n: int = CONFIRMACIONES_REQUERIDAS):
        self.n        = n
        self.conteos  = {}   # { nombre: int }

    def confirmar(self, nombre: str) -> bool:
        """
        Incrementa el contador del nombre detectado y resetea los demás.
        Retorna True cuando se alcanza el número de confirmaciones.
        """
        if nombre == "Desconocido":
            self.conteos = {}
            return False

        # Resetear todos los demás
        for k in list(self.conteos.keys()):
            if k != nombre:
                self.conteos[k] = 0

        self.conteos[nombre] = self.conteos.get(nombre, 0) + 1
        actual = self.conteos[nombre]

        print(f"  [{nombre}] confirmaciones: {actual}/{self.n}")

        if actual >= self.n:
            self.conteos[nombre] = 0   # reset para no volver a registrar
            return True
        return False

    def resetear(self, nombre: str):
        """Resetea el contador de un alumno específico."""
        self.conteos[nombre] = 0
